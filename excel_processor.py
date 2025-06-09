from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os
from datetime import datetime

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

def process_excel(filepath, original_filename):
    print("📄 엑셀 최적화 시작")

    try:
        wb = load_workbook(filepath)
        print("📂 엑셀 파일 열기 성공")
    except Exception as e:
        print(f"🚨 엑셀 파일 열기 실패: {e}")
        raise e

    ws = wb.active
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    for row in range(2, ws.max_row + 1):
        try:
            print(f"🔍 {row}번째 줄 처리 중")
            i_value = ws.cell(row=row, column=9).value
            q_cell = ws.cell(row=row, column=17)

            if i_value:
                # Q열 자동 분류 (기존 그대로)
                if "BPH-011" in i_value and "설치키트" in i_value and "단품" not in i_value:
                    q_cell.value = "sm_k_011"
                elif "BPH-041" in i_value and "설치키트" in i_value and "단품" not in i_value:
                    q_cell.value = "sm_k_041"
                elif "원수공급밸브" in i_value and "단품" in i_value:
                    q_cell.value = "sm_valve"
                elif "5M 튜빙선" in i_value and "단품" in i_value:
                    q_cell.value = "sm_5m"
                elif "공기청정기" in i_value or "BAS-017" in i_value:
                    q_cell.value = "bg_v1.0"
                elif (
                    "BPH-041" in i_value and
                    any(kw in i_value for kw in ["여분필터", "여분 필터", "여분필터 추가", "여분 필터 추가"])
                ):
                    q_cell.value = "4wb_v1.5"
                elif (
                    "BPH-041" in i_value and
                    not any(kw in i_value for kw in ["여분필터", "여분 필터", "여분필터 추가", "여분 필터 추가"])
                ):
                    q_cell.value = "4wa_v1.5"
                elif (
                    "BPH-011" in i_value and
                    ("블랙" in i_value or "Black" in i_value) and
                    any(kw in i_value for kw in ["여분필터", "여분 필터", "여분필터 추가", "여분 필터 추가"])
                ):
                    q_cell.value = "kb_v1.5"
                elif (
                    "BPH-011" in i_value and
                    ("블랙" in i_value or "Black" in i_value)
                ):
                    q_cell.value = "ka_v1.5"
                elif (
                    "BPH-011" in i_value and
                    ("화이트" in i_value or "White" in i_value) and
                    any(kw in i_value for kw in ["여분필터", "여분 필터", "여분필터 추가", "여분 필터 추가"])
                ):
                    q_cell.value = "wb_v1.5"
                elif (
                    "BPH-011" in i_value and
                    ("화이트" in i_value or "White" in i_value)
                ):
                    q_cell.value = "wa_v1.5"
                elif "A타입" in i_value:
                    q_cell.value = "pa_v1.0"
                elif "B타입" in i_value:
                    q_cell.value = "pa_v1.5"
                elif "5M 튜빙선" in i_value:
                    q_cell.value = "sm_5m"
                else:
                    q_cell.value = "false"
                    q_cell.fill = red_fill

                # R열 자동 분류
                r_cell = ws.cell(row=row, column=18)
                if "설치요청" in i_value:
                    r_cell.value = "p1"
                elif "직접설치" in i_value:
                    r_cell.value = "u1"
                elif "공기청정기" in i_value and "필터" not in i_value:
                    r_cell.value = "u1"
                else:
                    r_cell.value = ""

                # N열 자동 분류 + O열 오류 처리
                n_cell = ws.cell(row=row, column=14)
                o_cell = ws.cell(row=row, column=15)
                n_value = str(n_cell.value) if n_cell.value else ""

                print(f"🔎 N열 원본값: '{n_value}'")  # 디버깅용 출력

                if n_value:
                    if "aimu(모바일)" in n_value:
                        n_cell.value = "m_ca"
                        o_cell.value = ""
                    elif "aimu(PC)" in n_value:
                        n_cell.value = "m_ca"
                        o_cell.value = ""
                    elif "PC몰" in n_value:
                        n_cell.value = "m_ca"
                        o_cell.value = ""
                    elif "네이버페이" in n_value:
                        n_cell.value = "m_cna"
                        o_cell.value = ""
                    elif "네이버쇼핑" in n_value:
                        n_cell.value = "m_cna"
                        o_cell.value = ""
                    elif "지오코리아" in n_value:
                        n_cell.value = "m_gi"
                        o_cell.value = ""
                    elif "니코보코" in n_value:
                        n_cell.value = "m_ni"
                        o_cell.value = ""
                    elif "오늘의집" in n_value:
                        n_cell.value = "m_o"
                        o_cell.value = ""
                    elif "쿠팡" in n_value:
                        n_cell.value = "m_cp"
                        o_cell.value = ""
                    elif "스마트스토어" in n_value:
                        n_cell.value = "m_sm"
                        o_cell.value = ""
                    elif "플린트" in n_value:
                        n_cell.value = "m_pl"
                        o_cell.value = ""
                    elif "전화주문" in n_value:
                        n_cell.value = "m_or"
                        o_cell.value = ""
                    elif "샘플(무상)" in n_value:
                        n_cell.value = "m_sp_f"
                        o_cell.value = ""
                    elif "샘플(유상)" in n_value:
                        n_cell.value = "m_sp_p"
                        o_cell.value = ""
                    elif "직원구매" in n_value:
                        n_cell.value = "m_ep"
                        o_cell.value = ""
                    else:
                        o_cell.value = "false"
                        o_cell.fill = red_fill

        except Exception as e:
            print(f"🔥 {row}번째 줄 처리 중 오류 발생: {e}")
            ws.cell(row=row, column=17).value = "false"
            ws.cell(row=row, column=17).fill = red_fill

    today = datetime.today().strftime("%Y%m%d")
    new_filename = f"aimu_{today}.xlsx"
    save_path = os.path.join(BASE_DIR, "optimized_files", new_filename)
    wb.save(save_path)
    print(f"✅ 엑셀 최적화 완료 → {new_filename}")
    return save_path
